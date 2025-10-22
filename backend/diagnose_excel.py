"""
Скрипт для диагностики Excel файлов с матчами.
Показывает какие строки будут пропущены и почему.
"""
import openpyxl
import sys
import os

def diagnose_excel_file(filepath):
    """Диагностирует Excel файл и показывает проблемные строки"""
    
    if not os.path.exists(filepath):
        print(f"❌ Файл не найден: {filepath}")
        return
    
    print("="*80)
    print(f"📊 ДИАГНОСТИКА EXCEL ФАЙЛА: {os.path.basename(filepath)}")
    print("="*80)
    
    # Открываем Excel файл
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Получаем заголовки (первая строка)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value))
    
    print(f"\n📋 Найденные колонки ({len(headers)}):")
    for i, header in enumerate(headers, 1):
        print(f"  {i}. {header}")
    
    # Проверяем наличие обязательных колонок
    required_columns = {
        'дата': ['Дата', 'дата', 'Date'],
        'игрок_1': ['Игрок 1', 'игрок 1', 'Игрок1', 'Player 1'],
        'игрок_2': ['Игрок 2', 'игрок 2', 'Игрок2', 'Player 2'],
    }
    
    # Опциональные но желательные колонки
    optional_columns = {
        'счёт_новый': ['Счет матча игрока 1', 'Счет матча игрока 2'],
        'счёт_старый': ['Счёт', 'счёт', 'Счет', 'Score'],
        'рейтинг_1': ['Рейтинг игрока 1'],
        'рейтинг_2': ['Рейтинг игрока 2'],
        'стадия': ['Стадия', 'стадия', 'Этап', 'Stage'],
        'турнир': ['Турнир', 'турнир', 'Tournament']
    }
    
    print(f"\n🔍 Проверка обязательных колонок:")
    found_columns = {}
    for key, variants in required_columns.items():
        found = False
        for variant in variants:
            if variant in headers:
                found_columns[key] = variant
                print(f"  ✅ {key}: найдена как '{variant}'")
                found = True
                break
        if not found:
            print(f"  ❌ {key}: НЕ НАЙДЕНА! Ожидается одна из: {', '.join(variants)}")
    
    print(f"\n📋 Проверка опциональных колонок:")
    for key, variants in optional_columns.items():
        found = False
        for variant in variants:
            if variant in headers:
                if key.startswith('счёт_'):
                    # Для счёта проверяем обе колонки (новый формат)
                    if all(v in headers for v in variants):
                        found_columns[key] = variants
                        print(f"  ✅ {key}: найдены колонки {variants}")
                        found = True
                        break
                    elif variant in headers:
                        found_columns[key] = variant
                        print(f"  ✅ {key}: найдена как '{variant}'")
                        found = True
                        break
                else:
                    found_columns[key] = variant
                    print(f"  ✅ {key}: найдена как '{variant}'")
                    found = True
                    break
        if not found:
            print(f"  ⚠️  {key}: не найдена (опциональная)")
    
    # Определяем формат счёта
    has_new_score = 'счёт_новый' in found_columns
    has_old_score = 'счёт_старый' in found_columns
    
    if has_new_score:
        print(f"\n📊 Формат файла: НОВЫЙ (с детальными данными)")
        print(f"   Счёт из колонок: 'Счет матча игрока 1' и 'Счет матча игрока 2'")
    elif has_old_score:
        print(f"\n📊 Формат файла: СТАРЫЙ (упрощённый)")
        print(f"   Счёт из колонки: '{found_columns['счёт_старый']}'")
    else:
        print(f"\n⚠️  ВНИМАНИЕ: Не найдены колонки для счёта матча!")
        print(f"   Ожидается либо: 'Счет матча игрока 1' + 'Счет матча игрока 2'")
        print(f"   Либо: 'Счёт' (или 'Счет', 'Score')")
    
    
    # Анализируем каждую строку данных (начиная со второй)
    total_rows = ws.max_row - 1  # -1 потому что первая строка - заголовки
    valid_rows = 0
    invalid_rows = 0
    issues = []
    
    print(f"\n📈 Анализ строк данных (всего: {total_rows}):")
    print("-"*80)
    
    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        for col_num, header in enumerate(headers, 1):
            cell_value = ws.cell(row_num, col_num).value
            row_data[header] = cell_value
        
        # Находим значения по колонкам
        def get_value(key):
            if key not in found_columns:
                return None
            col_name = found_columns[key]
            return row_data.get(col_name)
        
        player1 = get_value('игрок_1')
        player2 = get_value('игрок_2')
        
        # Получаем счёт (новый или старый формат)
        score = None
        if has_new_score:
            score1 = ws.cell(row_num, headers.index(found_columns['счёт_новый'][0]) + 1).value
            score2 = ws.cell(row_num, headers.index(found_columns['счёт_новый'][1]) + 1).value
            if score1 is not None and score2 is not None:
                score = f"{score1}:{score2}"
        elif has_old_score:
            score = get_value('счёт_старый')
        
        # Проверяем валидность строки
        row_issues = []
        
        if not player1 or str(player1).strip() == '':
            row_issues.append("Отсутствует Игрок 1")
        
        if not player2 or str(player2).strip() == '':
            row_issues.append("Отсутствует Игрок 2")
        
        if not score or str(score).strip() == '':
            row_issues.append("Отсутствует Счёт")
        
        if row_issues:
            invalid_rows += 1
            issue_text = f"Строка {row_num}: {', '.join(row_issues)}"
            issues.append(issue_text)
            print(f"  ❌ {issue_text}")
            
            # Показываем содержимое проблемной строки
            print(f"     Данные: Игрок1='{player1}', Игрок2='{player2}', Счёт='{score}'")
        else:
            valid_rows += 1
            # Извлекаем имена и рейтинги
            def parse_player(player_str):
                if not player_str:
                    return None, None
                player_str = str(player_str)
                import re
                rating_match = re.search(r'rating:\s*(\d+(?:\.\d+)?)', player_str)
                rating = rating_match.group(1) if rating_match else "НЕТ"
                name = re.sub(r'\s*rating:\s*\d+(?:\.\d+)?', '', player_str).strip()
                return name, rating
            
            p1_name, p1_rating = parse_player(player1)
            p2_name, p2_rating = parse_player(player2)
            
            print(f"  ✅ Строка {row_num}: {p1_name} ({p1_rating}) vs {p2_name} ({p2_rating}) - {score}")
    
    # Итоговая статистика
    print("\n" + "="*80)
    print("📊 ИТОГОВАЯ СТАТИСТИКА:")
    print("="*80)
    print(f"Всего строк данных: {total_rows}")
    print(f"✅ Валидных строк: {valid_rows} ({valid_rows/total_rows*100:.1f}%)")
    print(f"❌ Невалидных строк: {invalid_rows} ({invalid_rows/total_rows*100:.1f}%)")
    
    if invalid_rows > 0:
        print(f"\n⚠️ ВНИМАНИЕ: {invalid_rows} строк будут ПРОПУЩЕНЫ при анализе!")
        print("\nПроблемные строки:")
        for issue in issues:
            print(f"  - {issue}")
    else:
        print(f"\n✅ Все строки валидны и будут обработаны!")
    
    print("\n" + "="*80)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование: python diagnose_excel.py <путь_к_файлу.xlsx>")
        print("\nПример:")
        print("  python diagnose_excel.py matches.xlsx")
        print("  python diagnose_excel.py C:\\Users\\User\\Desktop\\matches.xlsx")
    else:
        filepath = sys.argv[1]
        diagnose_excel_file(filepath)
